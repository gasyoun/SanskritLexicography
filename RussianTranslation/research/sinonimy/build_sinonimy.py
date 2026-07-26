"""Digitize V.V. Leonchenko's Sinonimy xlsx workbooks (VisualDCS/derived-data/Sinonimy)
into sinonimy.jsonl -- a per-sense synonym-evidence lane for B2 (see
../ROADMAP_ACL_LESSONS_2026.md, Wave 1: "Sinonimy xlsx->jsonl digitization").

Source-file dedup finding (H1491): the folder's four named groups (Глагольные
синонимы, Значения, S_P_D_F, Works-Share-Syn) reduce to THREE distinct datasets,
not four -- S_P_D_F/Глагольные синонимы.xlsx and Works-Share-Syn/data28.xlsx are
byte/row-identical copies of the top-level verb-synonyms workbook; S_P_D_F/
Синонимы существительных.xlsx and Works-Share-Syn/data29.xlsx are row-identical
copies of "Поиск синонимов..."; Works-Share-Syn/data30.xlsx is a structural
duplicate of Значения.xlsx. Only the three top-level canonical files are read.
Подобие по векторам.xlsx (a different, unvalidated vector-similarity method),
the frequency baseline, and the syntagmatic collocation table are out of scope
(not synonym-group/pair data) -- see README.md caveats.

Usage:  python build_sinonimy.py [path-to-VisualDCS]
"""
import json
import os
import sys

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

DEFAULT_VISUALDCS = os.path.normpath(os.path.join(
    os.path.dirname(os.path.abspath(__file__)), '..', '..', '..', '..', 'VisualDCS'))

OUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'sinonimy.jsonl')

MEANINGS_FILE = 'Значения.xlsx'
VERB_FILE = 'Глагольные синонимы_,без ограничений (2).xlsx'
SEARCH_FILE = 'Поиск синонимов в Цифровом корпусе Санскрита.xlsx'


def clean_lemma(cell):
    """Strip the |pipe| / /slash/ delimiters Leonchenko's export wraps lemmas in."""
    if cell is None:
        return None
    s = str(cell).strip()
    for ch in ('|', '/'):
        if s.startswith(ch) and s.endswith(ch) and len(s) > 1:
            s = s[1:-1]
    return s or None


def row_values(ws, min_row=2):
    for row in ws.iter_rows(min_row=min_row, values_only=True):
        # Trim the wide-dimension trailing Nones openpyxl reports for these sheets.
        vals = list(row)
        while vals and vals[-1] is None:
            vals.pop()
        if vals:
            yield vals


def build_sense_inventory(visualdcs_dir, rows):
    path = os.path.join(visualdcs_dir, 'derived-data', 'Sinonimy', MEANINGS_FILE)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    provenance = {'file': MEANINGS_FILE, 'sheet': 'Значения'}
    for vals in row_values(wb['Значения']):
        lemma, n_senses, *senses = vals
        if not lemma:
            continue
        rows.append({
            'type': 'sense_inventory',
            'source': 'leonchenko_sinonimy',
            'lemma': clean_lemma(lemma),
            'n_senses': n_senses,
            'senses': [s for s in senses if s],
            'provenance': provenance,
        })
    provenance = {'file': MEANINGS_FILE, 'sheet': 'Алфавитный порядок'}
    for vals in row_values(wb['Алфавитный порядок']):
        lemma, depth, gloss_anchor, *members = vals
        if not lemma:
            continue
        rows.append({
            'type': 'synonym_group_lemma',
            'source': 'leonchenko_sinonimy',
            'lemma': clean_lemma(lemma),
            'depth': depth,
            'gloss_anchor': gloss_anchor,
            'members': [clean_lemma(m) for m in members if m],
            'provenance': provenance,
        })
    wb.close()


def build_gloss_rings(visualdcs_dir, filename, sheet, pos, rows):
    path = os.path.join(visualdcs_dir, 'derived-data', 'Sinonimy', filename)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    provenance = {'file': filename, 'sheet': sheet}
    for vals in row_values(wb[sheet]):
        n_members, gloss, *members = vals
        if not gloss:
            continue
        rows.append({
            'type': 'synonym_group_gloss',
            'source': 'leonchenko_sinonimy',
            'pos': pos,
            'gloss': gloss,
            'n_members': n_members,
            'members': [clean_lemma(m) for m in members if m],
            'provenance': provenance,
        })
    wb.close()


def main():
    visualdcs_dir = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_VISUALDCS
    if not os.path.isdir(visualdcs_dir):
        sys.exit(f"VisualDCS dir not found: {visualdcs_dir}")

    rows = []
    build_sense_inventory(visualdcs_dir, rows)
    build_gloss_rings(visualdcs_dir, VERB_FILE, 'По дефинициям', 'verb', rows)
    build_gloss_rings(visualdcs_dir, SEARCH_FILE, 'По дефинициям', 'noun_or_general', rows)

    with open(OUT_PATH, 'w', encoding='utf-8') as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + '\n')

    by_type = {}
    for row in rows:
        by_type[row['type']] = by_type.get(row['type'], 0) + 1
    print(f"Wrote {len(rows)} rows to {OUT_PATH}")
    for t, n in sorted(by_type.items()):
        print(f"  {t}: {n}")


if __name__ == '__main__':
    main()
